import sys

try:
    from datetime import datetime
    import os
    import openpyxl
    from openpyxl.styles import Font
    from colorama import Fore, Style
    import sqlite3
except ImportError as e:
    print(Fore.RED + "Import error appeared. Reason: {}".format(e) + Style.RESET_ALL)
    sys.exit()

sys.path.append('service')

import crawl_processor as cp
import dorking_processor as dp
import networking_processor as np
import db_processing as db
import files_processing as fp

def create_report(short_domain, url, case_comment, report_file_type):
    try:
        ctime = datetime.now().strftime('%Y-%m-%d_%Hh%Mm%Ss')
        casename = short_domain.replace(".", "") + '_' + ctime + '.xlsx'
        foldername = short_domain.replace(".", "") + '_' + ctime
        db_casename = short_domain.replace(".", "")
        now = datetime.now()
        db_creation_date = str(now.year) + str(now.month) + str(now.day)
        report_folder = "report_{}".format(foldername)
        robots_filepath = report_folder + '//01-robots.txt'
        sitemap_filepath = report_folder + '//02-sitemap.txt'
        sitemap_links_filepath = report_folder + '//03-sitemap_links.txt'
        os.makedirs(report_folder, exist_ok=True)
        wb = openpyxl.Workbook()
        sheet_names = [
            "GENERAL INFO",
            "WHOIS",
            "SOCIAL MEDIAS",
            "SUBDOMAINS",
            "DNS SCAN",
            "SSL CERTIFICATE",
            "INTERNETDB SEARCH",
            "WEBSITE TECHNOLOGIES",
            "SITEMAP LINKS",
            "DORKING RESULTS"
        ]
        sheet = wb.active
        sheet.title = sheet_names[0]
        for name in sheet_names[1:]:
            wb.create_sheet(title=name)
        bold_font = Font(bold=True)

        print(Fore.GREEN + "Started scanning domain" + Style.RESET_ALL)
        print(Fore.GREEN + "Getting domain IP address" + Style.RESET_ALL)
        ip = cp.ip_gather(short_domain)
        print(Fore.GREEN + 'Gathering WHOIS information' + Style.RESET_ALL)
        res = cp.whois_gather(short_domain)
        print(Fore.GREEN + 'Processing e-mails gathering' + Style.RESET_ALL)
        mails = cp.mail_gather(url)
        print(Fore.GREEN + 'Processing subdomain gathering' + Style.RESET_ALL)
        subdomains, subdomains_amount = cp.subdomains_gather(url, short_domain)
        print(Fore.GREEN + 'Processing social medias gathering' + Style.RESET_ALL)
        social_medias = cp.sm_gather(url)
        print(Fore.GREEN + 'Processing subdomain analysis' + Style.RESET_ALL)
        subdomain_urls, subdomain_mails, subdomain_ip, sd_socials = cp.domains_reverse_research(subdomains, report_file_type)
        print(Fore.GREEN + 'Processing SSL certificate gathering' + Style.RESET_ALL)
        issuer, subject, notBefore, notAfter, commonName, serialNumber = np.get_ssl_certificate(short_domain)
        print(Fore.GREEN + 'Processing MX records gathering' + Style.RESET_ALL)
        mx_records = np.get_dns_info(short_domain)
        print(Fore.GREEN + 'Extracting robots.txt and sitemap.xml' + Style.RESET_ALL)
        robots_txt_result = np.get_robots_txt(short_domain, robots_filepath)
        sitemap_xml_result = np.get_sitemap_xml(short_domain, sitemap_filepath)
        sitemap_links_status, parsed_links = np.extract_links_from_sitemap(sitemap_links_filepath, sitemap_filepath, 'xlsx')
        print(Fore.GREEN + 'Gathering info about website technologies' + Style.RESET_ALL)
        web_servers, cms, programming_languages, web_frameworks, analytics, javascript_frameworks = np.get_technologies(url)
        print(Fore.GREEN + 'Processing Shodan InternetDB search' + Style.RESET_ALL)
        ports, hostnames, cpes, tags, vulns = np.query_internetdb(ip, report_file_type)
        print(Fore.GREEN + 'Processing Google Dorking' + Style.RESET_ALL)
        dorking_status, dorking_results = dp.transfer_results_to_xlsx(dp.get_dorking_query(short_domain))
        print(Fore.GREEN + 'Processing XLSX report for {} case...'.format(short_domain) + Style.RESET_ALL)
        common_socials = {key: social_medias.get(key, []) + sd_socials.get(key, []) for key in set(social_medias) | set(sd_socials)}
        for key in common_socials:
            common_socials[key] = list(set(common_socials[key]))
        total_socials = sum(len(values) for values in common_socials.values())

        ws = wb['GENERAL INFO']
        for col in ['1', '2', '3', '4', '5', '6', '7']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'SUBDOMAINS FOUND'
        ws['A2'] = 'SOCIAL MEDIAS FOUND'
        ws['A3'] = 'ROBOTS EXTRACTED?'
        ws['A4'] = 'SITEMAP.XML EXTRACTED?'
        ws['A5'] = 'SITEMAP.XML LINKS EXTRACTED?'
        ws['A6'] = 'DORKING STATUS'
        ws['A7'] = 'REPORT CREATION TIME'
        ws['B1'] = subdomains_amount
        ws['B2'] = total_socials
        ws['B3'] = robots_txt_result
        ws['B4'] = sitemap_xml_result
        ws['B5'] = sitemap_links_status
        ws['B6'] = dorking_status
        ws['B7'] = ctime

        ws = wb['WHOIS']
        for col in ['1', '2', '3', '4', '5', '6', '7', '8']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'SHORT DOMAIN'
        ws['A2'] = 'URL'
        ws['A3'] = 'IP ADDRESS'
        ws['A4'] = 'REGISTRAR'
        ws['A5'] = 'CREATION DATE'
        ws['A6'] = 'EXPIRATION DATE'
        ws['A7'] = 'NAME SERVERS'
        ws['A8'] = 'ORGANIZATION NAME'
        ws['B1'] = short_domain
        ws['B2'] = url
        ws['B3'] = ip
        ws['B4'] = res['registrar']
        ws['B5'] = res['creation_date']
        ws['B6'] = res['expiration_date']
        ws['B7'] = ', '.join(res['name_servers'])
        ws['B8'] = res['org']

        ws = wb['SOCIAL MEDIAS']
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            cell = f"{col}1"
            ws[cell].font = bold_font
            ws.column_dimensions[col].width = 70
        tw_links = common_socials['Twitter']
        inst_links = common_socials['Instagram']
        tg_links = common_socials['Telegram']
        tt_links = common_socials['TikTok']
        li_links = common_socials['LinkedIn']
        vk_links = common_socials['VKontakte']
        yt_links = common_socials['YouTube']
        wc_links = common_socials['WeChat']
        ok_links = common_socials['Odnoklassniki']
        fb_links = common_socials['Facebook']
        ws['A1'] = 'FACEBOOK'
        ws['B1'] = 'TWITTER'
        ws['C1'] = 'INSTAGRAM'
        ws['D1'] = 'TELEGRAM'
        ws['E1'] = 'TIKTOK'
        ws['F1'] = 'LINKEDIN'
        ws['G1'] = 'VKONTAKTE'
        ws['H1'] = 'YOUTUBE'
        ws['I1'] = 'ODNOKLASSNIKI'
        ws['J1'] = 'WECHAT'

        for i in range(len(fb_links)):
            ws[f"A{i + 2}"] = fb_links[i]
        for i in range(len(tw_links)):
            ws[f"B{i + 2}"] = tw_links[i]
        for i in range(len(inst_links)):
            ws[f"C{i + 2}"] = inst_links[i]
        for i in range(len(tg_links)):
            ws[f"D{i + 2}"] = tg_links[i]
        for i in range(len(tt_links)):
            ws[f"E{i + 2}"] = tt_links[i]
        for i in range(len(li_links)):
            ws[f"F{i + 2}"] = li_links[i]
        for i in range(len(vk_links)):
            ws[f"G{i + 2}"] = vk_links[i]
        for i in range(len(yt_links)):
            ws[f"H{i + 2}"] = yt_links[i]
        for i in range(len(ok_links)):
            ws[f"I{i + 2}"] = ok_links[i]
        for i in range(len(wc_links)):
            ws[f"J{i + 2}"] = wc_links[i]

        ws = wb['SUBDOMAINS']
        for col in ['A', 'B', 'C']:
            cell = f"{col}1"
            ws[cell].font = bold_font
            ws.column_dimensions[col].width = 70
        ws['A1'] = 'FOUNDED SUBDOMAINS'
        ws['B1'] = 'SUBDOMAIN IP ADDRESSES (NOT CORRELATED)'
        ws['C1'] = 'SUBDOMAIN EMAILS (NOT CORRELATED)'
        try:
            for i in range(len(subdomain_urls)):
                ws[f"A{i + 2}"] = str(subdomain_urls[i])
            for i in range(len(subdomain_ip)):
                ws[f"B{i + 2}"] = str(subdomain_ip[i])
            for i in range(len(subdomain_mails)):
                ws[f"C{i + 2}"] = str(subdomain_mails[i])
        except Exception as e:
            print(Fore.RED + "Error appeared when writing some information about subdomains in XLSX file. Reason: {}".format(e))
            pass

        ws = wb['DNS SCAN']
        for col in ['1', '2']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'NAME SERVERS'
        ws['A2'] = 'MX ADDRESSES'
        ws['B1'] = ', '.join(res['name_servers'])
        ws['B2'] = mx_records

        ws = wb['SSL CERTIFICATE']
        for col in ['1', '2', '3', '4', '5', '6']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'ISSUER'
        ws['A2'] = 'SUBJECT'
        ws['A3'] = 'NOT BEFORE'
        ws['A4'] = 'NOT AFTER'
        ws['A5'] = 'CERTIFICATE NAME'
        ws['A6'] = 'CERTIFICATE SERIAL NUMBER'
        ws['B1'] = issuer
        ws['B2'] = subject
        ws['B3'] = notBefore
        ws['B4'] = notAfter
        ws['B5'] = commonName
        ws['B6'] = serialNumber

        ws = wb['INTERNETDB SEARCH']
        for col in ['1', '2', '3', '4']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws['I1'].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'OPEN PORTS'
        ws['A2'] = 'HOSTNAMES'
        ws['A3'] = 'TAGS'
        ws['A4'] = 'CPEs'
        ws['I1'] = 'POTENTIAL VULNERABILITIES'
        ws['B1'] = str(ports)
        ws['B2'] = str(hostnames)
        ws['B3'] = str(tags)
        ws['B4'] = str(cpes)
        for i in range(len(vulns)):
            ws[f"I{i + 2}"] = str(vulns[i])

        ws = wb['WEBSITE TECHNOLOGIES']
        for col in ['1', '2', '3', '4', '5', '6']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'WEB SERVERS'
        ws['A2'] = 'CMS'
        ws['A3'] = 'USED PROGRAMMING LANGUAGES'
        ws['A4'] = 'USED WEB FRAMEWORKS'
        ws['A5'] = 'ANALYTICS SERVICE'
        ws['A6'] = 'USED JAVASCRIPT FRAMEWORKS'
        ws['B1'] = str(web_servers)
        ws['B2'] = str(cms)
        ws['B3'] = str(programming_languages)
        ws['B4'] = str(web_frameworks)
        ws['B5'] = str(analytics)
        ws['B6'] = str(javascript_frameworks)

        ws = wb['SITEMAP LINKS']
        ws.column_dimensions['A'].width = 80
        for i in range(len(parsed_links)):
            ws[f"A{i + 1}"] = str(parsed_links[i])

        ws = wb['DORKING RESULTS']
        ws.column_dimensions['A'].width = 80
        for i in range(len(dorking_results)):
            ws[f"A{i + 1}"] = str(dorking_results[i])

        report_file = report_folder + "//" + casename
        wb.save(report_file)
        print(Fore.GREEN + "Report for {} case was created at {}".format(''.join(short_domain), ctime) + Style.RESET_ALL)
        robots_content, sitemap_content, sitemap_links_content, dorking_content = fp.get_db_columns(report_folder)
        xlsx_blob = fp.get_blob(report_file)
        db.insert_blob('XLSX', xlsx_blob, db_casename, db_creation_date, case_comment, robots_content, sitemap_content, sitemap_links_content, dorking_content)
    except Exception as e:
        print(Fore.RED + 'Unable to create XLSX report. Reason: {}'.format(e))