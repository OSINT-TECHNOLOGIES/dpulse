import sys

sys.path.append('service')

from logs_processing import logging
import db_processing as db
import files_processing as fp

try:
    from datetime import datetime
    import os
    import openpyxl
    from openpyxl.styles import Font
    from colorama import Fore, Style
    import sqlite3
    from urllib.parse import unquote
except ImportError as e:
    print(Fore.RED + "Import error appeared. Reason: {}".format(e) + Style.RESET_ALL)
    sys.exit()

def create_report(short_domain, url, case_comment, data_array, report_info_array, pagesearch_ui_mark, pagesearch_keyword, end):
    try:
        ip = data_array[0]
        res = data_array[1]
        subdomain_urls = data_array[3]
        subdomains_amount = data_array[4]
        subdomain_mails = data_array[6]
        subdomain_ip = data_array[8]
        issuer = data_array[9]
        subject = data_array[10]
        notBefore = data_array[11]
        notAfter = data_array[12]
        commonName = data_array[13]
        serialNumber = data_array[14]
        mx_records = data_array[15]
        robots_txt_result = data_array[16]
        sitemap_xml_result = data_array[17]
        sitemap_links_status = data_array[18]
        web_servers = data_array[19]
        cms = data_array[20]
        programming_languages = data_array[21]
        web_frameworks = data_array[22]
        analytics = data_array[23]
        javascript_frameworks = data_array[24]
        ports = data_array[25]
        hostnames = data_array[26]
        cpes = data_array[27]
        tags = data_array[28]
        vulns = data_array[29]
        dorking_status = data_array[42]
        common_socials = data_array[30]
        total_socials = data_array[31]
        ps_emails_return = data_array[32]
        accessible_subdomains = data_array[33]
        emails_amount = data_array[34]
        files_counter = data_array[35]
        cookies_counter = data_array[36]
        api_keys_counter = data_array[37]
        website_elements_counter = data_array[38]
        exposed_passwords_counter = data_array[39]
        total_links_counter = data_array[40]
        accessed_links_counter = data_array[41]
        cleaned_dorking = data_array[42]
        vt_cats = data_array[43]
        vt_deturls = data_array[44]
        vt_detsamples = data_array[45]
        vt_undetsamples = data_array[46]
        st_alexa = data_array[47]
        st_apex = data_array[48]
        st_hostname = data_array[49]
        st_alivesds = data_array[50]
        st_txt = data_array[51]
        a_records_list = data_array[52]
        mx_records_list = data_array[53]
        ns_records_list = data_array[54]
        soa_records_list = data_array[55]

        casename = report_info_array[0]
        db_casename = report_info_array[1]
        db_creation_date = report_info_array[2]
        report_folder = report_info_array[3]
        report_ctime = report_info_array[6]
        api_scan_db = report_info_array[7]
        used_api_flag = report_info_array[8]
        os.makedirs(report_folder, exist_ok=True)

        if '2' in used_api_flag:
            st_a_combined = []
            if len(a_records_list) > 0:
                if len(a_records_list) == 1:
                    record = a_records_list[0]
                    st_a_combined = [f"IPv4 address: {record.get('ip', '')}, owned by {record.get('organization', '')}"]
                else:
                    st_a_combined = [f"IPv4 address: {record.get('ip', '')}, owned by {record.get('organization', '')}" for record in a_records_list]

            st_mx_combined = []
            if len(mx_records_list) > 0:
                if len(mx_records_list) == 1:
                    record = mx_records_list[0]
                    st_mx_combined = [f"Hostname {record.get('mx_hostname', '')} with priority={record.get('mx_priority', '')}, owned by {record.get('mx_organization', '')}"]
                else:
                    st_mx_combined = [f"Hostname {record.get('mx_hostname', '')} with priority={record.get('mx_priority', '')}, owned by {record.get('mx_organization', '')}" for record in mx_records_list]

            st_ns_combined = []
            if len(ns_records_list) > 0:
                if len(ns_records_list) == 1:
                    record = ns_records_list[0]
                    st_ns_combined = [f"Nameserver: {record.get('ns_nameserver', '')}, owned by {record.get('ns_organization', '')}"]
                else:
                    st_ns_combined = [f"Nameserver: {record.get('ns_nameserver', '')}, owned by {record.get('ns_organization', '')}" for record in ns_records_list]

            st_soa_combined = []
            if len(soa_records_list) > 0:
                if len(soa_records_list) == 1:
                    record = soa_records_list[0]
                    st_soa_combined = [f"Email: {record.get('soa_email', '')}, TTL={record.get('soa_ttl', '')}"]
                else:
                    st_soa_combined = [f"Email: {record.get('soa_email', '')}, TTL={record.get('soa_ttl', '')}" for record in soa_records_list]
        else:
            st_soa_combined = st_ns_combined = st_mx_combined = st_a_combined = st_txt = st_alivesds = ['No results because user did not selected SecurityTrails API scan']

        if len(ps_emails_return) > 0:
            subdomain_mails += ps_emails_return
            subdomain_mails = list(set(subdomain_mails))
            subdomain_mails_cleaned = []
            substrings = ['m=Base64', 'Ë','Á','Æ','Å','Ä','Ò','Á','ó','ð','É','ë','â']
            for substring in substrings:
                if any(substring in s for s in subdomain_mails):
                    subdomain_mails.remove(next(s for s in subdomain_mails if substring in s))
            for email in subdomain_mails:
                new_emails = email.split(', ')
                subdomain_mails_cleaned.extend(new_emails)
        else:
            subdomain_mails = list(set(subdomain_mails))
            subdomain_mails_cleaned = []
            substrings = ['m=Base64', 'Ë','Á','Æ','Å','Ä','Ò','Á','ó','ð','É','ë','â']
            for substring in substrings:
                if any(substring in s for s in subdomain_mails):
                    subdomain_mails.remove(next(s for s in subdomain_mails if substring in s))
            for email in subdomain_mails:
                new_emails = email.split(', ')
                subdomain_mails_cleaned.extend(new_emails)

        wb = openpyxl.Workbook()
        sheet_names = [
            "GENERAL INFO",
            "WHOIS",
            "SOCIAL MEDIAS",
            "SUBDOMAINS",
            "DNS & SSL",
            "WEB INFO",
            "PRE-PENTEST INFO",
            "DORKING RESULTS",
            "PAGESEARCH",
            "PAGESEARCH (SI)",
            "VIRUSTOTAL API",
            "SECURITYTRAILS API"
        ]
        sheet = wb.active
        sheet.title = sheet_names[0]
        for name in sheet_names[1:]:
            wb.create_sheet(title=name)
        bold_font = Font(bold=True)

        ws = wb['GENERAL INFO']
        for col in ['1', '2', '3', '4', '5', '6', '7', '8', '9']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'TARGET DOMAIN'
        ws['A2'] = 'TOTAL SUBDOMAINS FOUND'
        ws['A3'] = 'TOTAL SOCIAL MEDIAS LINKS FOUND'
        ws['A4'] = 'STATUS OF ROBOTS.TXT EXTRACTION'
        ws['A5'] = 'STATUS OF SITEMAP.XML EXTRACTION'
        ws['A6'] = 'STATUS OF SITEMAP.XML LINKS EXTRACTION'
        ws['A7'] = 'GOOGLE DORKING STATUS'
        ws['A8'] = 'PAGESEARCH CONDUCTION'
        ws['A9'] = 'REPORT CREATION TIME'
        ws['B1'] = short_domain
        ws['B2'] = subdomains_amount
        ws['B3'] = total_socials
        ws['B4'] = robots_txt_result
        ws['B5'] = sitemap_xml_result
        ws['B6'] = sitemap_links_status
        ws['B8'] = pagesearch_ui_mark
        ws['B9'] = report_ctime

        ws = wb['WHOIS']
        for col in ['1', '2', '3', '4', '5', '6', '7', '8']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'SHORT DOMAIN'
        ws['A2'] = 'URL'
        ws['A3'] = 'IP ADDRESS'
        ws['A4'] = 'REGISTRAR'
        ws['A5'] = 'CREATION DATE'
        ws['A6'] = 'EXPIRATION DATE'
        ws['A7'] = 'NAME SERVERS'
        ws['A8'] = 'ORGANIZATION NAME'
        ws['B1'] = short_domain
        ws['B2'] = url
        ws['B3'] = ip
        ws['B4'] = res['registrar']
        ws['B5'] = res['creation_date']
        ws['B6'] = res['expiration_date']
        ws['B7'] = ', '.join(res['name_servers'])
        ws['B8'] = res['org']

        ws = wb['SOCIAL MEDIAS']
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            cell = f"{col}1"
            ws[cell].font = bold_font
            ws.column_dimensions[col].width = 70
        tw_links = common_socials['Twitter']
        inst_links = common_socials['Instagram']
        tg_links = common_socials['Telegram']
        tt_links = common_socials['TikTok']
        li_links = common_socials['LinkedIn']
        vk_links = common_socials['VKontakte']
        yt_links = common_socials['YouTube']
        wc_links = common_socials['WeChat']
        ok_links = common_socials['Odnoklassniki']
        fb_links = common_socials['Facebook']
        ws['A1'] = 'FACEBOOK'
        ws['B1'] = 'TWITTER'
        ws['C1'] = 'INSTAGRAM'
        ws['D1'] = 'TELEGRAM'
        ws['E1'] = 'TIKTOK'
        ws['F1'] = 'LINKEDIN'
        ws['G1'] = 'VKONTAKTE'
        ws['H1'] = 'YOUTUBE'
        ws['I1'] = 'ODNOKLASSNIKI'
        ws['J1'] = 'WECHAT'

        for i in range(len(fb_links)):
            ws[f"A{i + 2}"] = fb_links[i]
        for i in range(len(tw_links)):
            ws[f"B{i + 2}"] = tw_links[i]
        for i in range(len(inst_links)):
            ws[f"C{i + 2}"] = inst_links[i]
        for i in range(len(tg_links)):
            ws[f"D{i + 2}"] = tg_links[i]
        for i in range(len(tt_links)):
            ws[f"E{i + 2}"] = tt_links[i]
        for i in range(len(li_links)):
            ws[f"F{i + 2}"] = li_links[i]
        for i in range(len(vk_links)):
            ws[f"G{i + 2}"] = vk_links[i]
        for i in range(len(yt_links)):
            ws[f"H{i + 2}"] = yt_links[i]
        for i in range(len(ok_links)):
            ws[f"I{i + 2}"] = ok_links[i]
        for i in range(len(wc_links)):
            ws[f"J{i + 2}"] = wc_links[i]

        ws = wb['SUBDOMAINS']
        ws['A1'] = 'FOUND SUBDOMAINS'
        ws['B1'] = 'SUBDOMAIN IP ADDRESSES (NOT CORRELATED)'
        ws['C1'] = 'SUBDOMAIN EMAILS (NOT CORRELATED)'
        for col in ['A', 'B', 'C']:
            cell = f"{col}1"
            ws[cell].font = bold_font
            ws.column_dimensions[col].width = 70
        try:
            for i in range(len(subdomain_urls)):
                ws[f"A{i + 2}"] = str(subdomain_urls[i])
            for i in range(len(subdomain_ip)):
                ws[f"B{i + 2}"] = str(subdomain_ip[i])
            for i in range(len(subdomain_mails)):
                ws[f"C{i + 2}"] = str(subdomain_mails[i])
        except Exception as e:
            print(Fore.RED + "Error appeared when writing some information about subdomains in XLSX file. See journal for details")
            logging.error(f'ERROR WHEN WRITING INFORMATION IN XLSX REPORT. REASON: {e}')
            pass

        ws = wb['DNS & SSL']
        for col in ['1', '2', '3', '4', '5', '6', '7', '8']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 60
        ws['A1'] = '(DNS) NAME SERVERS'
        ws['A2'] = '(DNS) MX ADDRESSES'
        ws['A3'] = '(SSL) ISSUER'
        ws['A4'] = '(SSL) SUBJECT'
        ws['A5'] = '(SSL) NOT BEFORE'
        ws['A6'] = '(SSL) NOT AFTER'
        ws['A7'] = '(SSL) CERTIFICATE NAME'
        ws['A8'] = '(SSL) CERTIFICATE SERIAL NUMBER'
        ws['B1'] = ', '.join(res['name_servers'])
        ws['B2'] = mx_records
        ws['B3'] = issuer
        ws['B4'] = subject
        ws['B5'] = notBefore
        ws['B6'] = notAfter
        ws['B7'] = commonName
        ws['B8'] = serialNumber

        ws = wb['WEB INFO']
        for col in ['1', '2', '3', '4', '5', '6', '7', '8']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'WEB SERVERS'
        ws['A2'] = 'CMS'
        ws['A3'] = 'USED PROGRAMMING LANGUAGES'
        ws['A4'] = 'USED WEB FRAMEWORKS'
        ws['A5'] = 'ANALYTICS SERVICE'
        ws['A6'] = 'USED JAVASCRIPT FRAMEWORKS'
        ws['A7'] = 'TAGS'
        ws['A8'] = 'CPEs'
        ws['B1'] = ', '.join(web_servers)
        ws['B2'] = ', '.join(cms)
        ws['B3'] = ', '.join(programming_languages)
        ws['B4'] = ', '.join(web_frameworks)
        ws['B5'] = ', '.join(analytics)
        ws['B6'] = ', '.join(javascript_frameworks)
        ws['B7'] = ', '.join(str(item) for item in tags)
        ws['B8'] = ', '.join(str(item) for item in cpes)

        ws = wb['PRE-PENTEST INFO']
        ws['A1'] = 'OPEN PORTS'
        ws['A2'] = 'HOSTNAMES'
        ws['F1'] = 'POTENTIAL VULNERABILITIES'
        for col in ['1', '2']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws['F1'].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['B1'] = ', '.join(str(item) for item in ports)
        ws['B2'] = ', '.join(str(item) for item in hostnames)
        for i in range(len(vulns)):
            ws[f"F{i + 2}"] = str(vulns[i])

        ws = wb['DORKING RESULTS']
        ws.column_dimensions['A'].width = 80
        for i, item in enumerate(cleaned_dorking, start=2):
            ws[f"A{i}"] = str(item)

        ws = wb['PAGESEARCH']
        for col in ['1', '2', '3', '4', '5', '6', '7']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'ACCESSIBLE SUBDOMAINS'
        ws['A2'] = 'ADDITIONAL EMAILS FOUND'
        ws['A3'] = 'FOUND DOCUMENTS'
        ws['A4'] = 'FOUND COOKIES'
        ws['A5'] = 'FOUND API KEYS'
        ws['A6'] = 'WEB ELEMENTS FOUND'
        ws['A7'] = 'FOUND EXPOSED PASSWORDS'
        ws['B1'] = accessible_subdomains
        ws['B2'] = emails_amount
        ws['B3'] = files_counter
        ws['B4'] = cookies_counter
        ws['B5'] = api_keys_counter
        ws['B6'] = website_elements_counter
        ws['B7'] = exposed_passwords_counter

        ws = wb['PAGESEARCH (SI)']
        for col in ['1', '2']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'TOTAL LINKS AMOUNT'
        ws['A2'] = 'AMOUNT OF ACCESSED LINKS'
        ws['B1'] = total_links_counter
        ws['B2'] = accessed_links_counter

        ws = wb['VIRUSTOTAL API']
        for col in ['1', '2', '3', '4']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 60
        ws['A1'] = 'CATEGORIES'
        ws['A2'] = 'DETECTED URLS'
        ws['A3'] = 'DETECTED SAMPLES'
        ws['A4'] = 'UNDETECTED SAMPLES'
        ws['B1'] = vt_cats
        ws['B2'] = vt_deturls
        ws['B3'] = vt_detsamples
        ws['B4'] = vt_undetsamples

        ws = wb['SECURITYTRAILS API']
        for col in ['1', '2', '3']:
            cell = f"A{col}"
            ws[cell].font = bold_font
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['E'].width = 70
        ws.column_dimensions['F'].width = 70
        ws.column_dimensions['G'].width = 70
        ws.column_dimensions['H'].width = 70
        ws.column_dimensions['I'].width = 70
        ws.column_dimensions['J'].width = 70
        ws['A1'] = 'ALEXA RANK'
        ws['A2'] = 'APEX DOMAIN'
        ws['A3'] = 'HOSTNAME'
        ws['E1'] = 'A RECORDS'
        ws['F1'] = 'MX RECORDS'
        ws['G1'] = 'NS RECORDS'
        ws['H1'] = 'SOA RECORDS'
        ws['I1'] = 'TXT RECORDS'
        ws['J1'] = 'SUBDOMAINS LIST'
        ws['E1'].font = ws['F1'].font = ws['G1'].font = ws['H1'].font = ws['I1'].font = ws['J1'].font = bold_font
        ws['B1'] = st_alexa
        ws['B2'] = st_apex
        ws['B3'] = st_hostname

        for i in range(len(st_a_combined)):
            ws[f"E{i + 2}"] = str(st_a_combined[i])
        for i in range(len(st_mx_combined)):
            ws[f"F{i + 2}"] = str(st_mx_combined[i])
        for i in range(len(st_ns_combined)):
            ws[f"G{i + 2}"] = str(st_ns_combined[i])
        for i in range(len(st_soa_combined)):
            ws[f"H{i + 2}"] = str(st_soa_combined[i])
        for i in range(len(st_txt)):
            ws[f"I{i + 2}"] = str(st_txt[i])
        for i in range(len(st_alivesds)):
            ws[f"J{i + 2}"] = str(st_alivesds[i])


        report_file = report_folder + "//" + casename
        wb.save(report_file)
        print(Fore.GREEN + "XLSX report for {} case was created at {}".format(''.join(short_domain), report_ctime) + Style.RESET_ALL)
        print(Fore.GREEN + f"Scan elapsed time: {end}" + Style.RESET_ALL)
        robots_content, sitemap_content, sitemap_links_content, dorking_content = fp.get_db_columns(report_folder)
        xlsx_blob = fp.get_blob(report_file)
        db.insert_blob('XLSX', xlsx_blob, db_casename, db_creation_date, case_comment, robots_content, sitemap_content, sitemap_links_content, dorking_content, api_scan_db)
    except Exception as e:
        print(Fore.RED + 'Unable to create XLSX report. See journal for details')
        logging.error(f'XLSX REPORT CREATION: ERROR. REASON: {e}')
